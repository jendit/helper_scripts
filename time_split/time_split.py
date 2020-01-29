import datetime
import os

import openpyxl

TIME_SPLIT_SHORTCUTS = 'time_split_shortcuts.xlsx'
TIME_SPLIT_RECORDS = 'time_split_records.xlsx'

SEP = os.path.sep
HOME = os.path.expanduser('~')

TIME_SPLIT_SHORTCUTS_PATH = HOME + SEP + 'Documents' + SEP + TIME_SPLIT_SHORTCUTS
TIME_SPLIT_RECORDS_PATH = HOME + SEP + 'Documents' + SEP + TIME_SPLIT_RECORDS


def read_shortcurts():
    ''' Read shortcut list file into a dictionary.
    
    return - A dictionary with the shortcuts and task name combinations.
    '''
    sc = {}
    wb = openpyxl.load_workbook(TIME_SPLIT_SHORTCUTS_PATH)
    sheet = wb.active
    row_modifyer = 0
    while True:
        cell_1 = sheet.cell(row = 2 + row_modifyer, column = 1)
        cell_2 = sheet.cell(row = 2 + row_modifyer, column = 2)
        if cell_1.value == None or cell_2.value == None:
            break
        sc[cell_1.value] = cell_2.value
        row_modifyer += 1
    wb.close()
    return sc


def set_task(task_name = ''):
    ''' Write a task to the Excel-Document.

    If the task_name is provided, a new task will be appended.
    Regardless if a task_name is provided, the last open task will be closed.

    task_name - optional paramenter with the name of the task.
    '''
    wb = openpyxl.load_workbook(TIME_SPLIT_RECORDS_PATH)
    sheet = wb.active
    row_count = len(sheet['A'])  # calculate the len by getting the first column as this one will always be filled

    last_end = sheet.cell(row = row_count, column = 3)
    if last_end.value == None:
        last_end.value = datetime.datetime.now()

    if task_name:
        new_row = [task_name, datetime.datetime.now()]
        sheet.append(new_row)

    wb.save(TIME_SPLIT_RECORDS_PATH)
    wb.close()


def input_time_split():
    '''Input dialog for the user.

    The dialog needs to get the shortcut of a task name to start a task.
    If a previous task is not finished, it will be closed automatically.
    If no shortcut for a task is provided, the last task in the list will be closed.
    '''
    print('Please provide a shortcut for a task (see >> time_split_shortcuts.xlsx << for reference).')
    print('Or leave empty to just close the last active task.')
    task_shortcut = input('Task: ')
    if task_shortcut:
        try:
            shortcut_list = read_shortcurts()
            set_task(shortcut_list[task_shortcut])
        except KeyError:
            print('ERROR, the given shortcut is unknown.')
    else:
        set_task()


if __name__ == '__main__':
    input_time_split()
